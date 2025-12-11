
from django.shortcuts import redirect, render
from django.http.response import HttpResponse
from rest_framework.response import Response
from rest_framework.decorators import api_view, authentication_classes, permission_classes, renderer_classes
from django.utils import timezone

from backend.vendor.serializers import VendorSerializers
from helpers.custom_functions import CustomPagination
from product.serializers import *
from product.models import *
import json
from django.db.models import Q
# from num2words import num2words
from django.db.models import Sum, IntegerField,FloatField
from django.db.models.functions import Cast
from decimal import Decimal, ROUND_DOWN
import math
from datetime import date, datetime
from time import strftime
from helpers.validations import hosturl
from django.utils.timezone import now
import locale
from dateutil.relativedelta import relativedelta
import random
from rest_framework import pagination
from rest_framework.generics import GenericAPIView



# Create your views here.
class product_list_pagination(GenericAPIView):
    pagination_class = CustomPagination
    def post(self, request):
        products_obj = Product.objects.filter(isActive=True).order_by('name')
        if products_obj.exists():
            page4 = self.paginate_queryset(products_obj)
            serializer=CustomProductSerializers(page4,many=True)
            return self.get_paginated_response(serializer.data)
        else:
            return self.get_paginated_response([])
        

@api_view(['GET'])
def product_list(request):
    products_obj = Product.objects.filter(isActive=True).order_by('name')
    if products_obj.exists():
        serializer=ProductSerializers(products_obj,many=True)
        return Response({"response":{"n":1,"msg":"Product list fetched successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":1,"msg":"No products found","status":"success"},"data":[]})
        


@api_view(['POST'])
def add_product(request):
    data = request.data.copy()
    data['isActive']=True
    print("data",data)
    alredy_exists = Product.objects.filter(Q(name__iexact=data.get('name')) & Q(isActive=True))
    if alredy_exists.exists():
        return Response({"response":{"n":0,"msg":"Product  with this name already exists","status":"error"}})
    serializer = ProductSerializers(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response({"response":{"n":1,"msg":"Product added successfully","status":"success"},"data":serializer.data})
    else:
        first_key, first_value = next(iter(serializer.errors.items()))
        return Response({"response":{"n":0,"msg":first_key+' : '+ first_value[0],"status":"error"},"errors":serializer.errors})
            

@api_view(['POST'])
def update_product(request):
    data = request.data.copy()
    product_id = data.get('id')
    try:
        product_obj = Product.objects.get(id=product_id, isActive=True)
    except Product.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product product not found","status":"error"}})
    
    alredy_exists = Product.objects.filter(Q(name__iexact=data.get('name')) & ~Q(id=product_id) & Q(isActive=True))
    if alredy_exists.exists():
        return Response({"response":{"n":0,"msg":"Product product with this name already exists","status":"error"}})
    
    serializer = ProductSerializers(product_obj, data=data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({"response":{"n":1,"msg":"Product product updated successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":0,"msg":"Invalid data","status":"error"},"errors":serializer.errors})
    

@api_view(['POST'])
def delete_product(request):
    data = request.data.copy()
    product_id = data.get('id')
    try:
        product_obj = Product.objects.get(id=product_id, isActive=True)
    except Product.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product product not found","status":"error"}})
    
    product_obj.isActive = False
    product_obj.save()
    return Response({"response":{"n":1,"msg":"Product product deleted successfully","status":"success"}})

@api_view(['POST'])
def product_by_id(request):

    data = request.data.copy()
    product_id = data.get('id')
    try:
        product_obj = Product.objects.get(id=product_id, isActive=True)
    except Product.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product product not found","status":"error"}})
    
    serializer=CustomProductSerializers(product_obj)
    return Response({"response":{"n":1,"msg":"Product product fetched successfully","status":"success"},"data":serializer.data})




@api_view(['GET'])
def product_category_list(request):
    categories_obj = ProductCategory.objects.filter(isActive=True).order_by('name')
    if categories_obj.exists():
        serializer=ProductCategorySerializers(categories_obj,many=True)
        return Response({"response":{"n":1,"msg":"Product category list fetched successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":1,"msg":"No product categories found","status":"success"},"data":[]})


class product_category_list_pagination(GenericAPIView):
    pagination_class = CustomPagination
    def post(self, request):
        productcategorys_obj = ProductCategory.objects.filter(isActive=True).order_by('name')
        if productcategorys_obj.exists():
            page4 = self.paginate_queryset(productcategorys_obj)
            serializer=ProductCategorySerializers(page4,many=True)
            return self.get_paginated_response(serializer.data)
        else:
            return self.get_paginated_response([])

@api_view(['POST'])
def add_product_category(request):
    data = request.data.copy()
    data['isActive']=True

    alredy_exists = ProductCategory.objects.filter(Q(name__iexact=data.get('name')) & Q(isActive=True))
    if alredy_exists.exists():
        return Response({"response":{"n":0,"msg":"Product category with this name already exists","status":"error"}})
    
    serializer = ProductCategorySerializers(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response({"response":{"n":1,"msg":"Product category added successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":0,"msg":"Invalid data","status":"error"},"errors":serializer.errors})
    
@api_view(['POST'])
def update_product_category(request):
    data = request.data.copy()
    category_id = data.get('id')
    try:
        category_obj = ProductCategory.objects.get(id=category_id, isActive=True)
    except ProductCategory.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product category not found","status":"error"}})
    
    alredy_exists = ProductCategory.objects.filter(Q(name__iexact=data.get('name')) & ~Q(id=category_id) & Q(isActive=True))
    if alredy_exists.exists():
        return Response({"response":{"n":0,"msg":"Product category with this name already exists","status":"error"}})
    
    serializer = ProductCategorySerializers(category_obj, data=data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({"response":{"n":1,"msg":"Product category updated successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":0,"msg":"Invalid data","status":"error"},"errors":serializer.errors})
    

@api_view(['POST'])
def delete_product_category(request):
    data = request.data.copy()
    category_id = data.get('id')
    try:
        category_obj = ProductCategory.objects.get(id=category_id, isActive=True)
    except ProductCategory.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product category not found","status":"error"}})
    
    category_obj.isActive = False
    category_obj.save()
    return Response({"response":{"n":1,"msg":"Product category deleted successfully","status":"success"}})

@api_view(['POST'])
def product_category_by_id(request):

    data = request.data.copy()
    category_id = data.get('id')
    try:
        category_obj = ProductCategory.objects.get(id=category_id, isActive=True)
    except ProductCategory.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product category not found","status":"error"}})
    
    serializer=ProductCategorySerializers(category_obj)
    return Response({"response":{"n":1,"msg":"Product category fetched successfully","status":"success"},"data":serializer.data})


class product_sub_category_list_pagination(GenericAPIView):
    pagination_class = CustomPagination
    def post(self, request):
        productsubcategorys_obj = ProductSubCategory.objects.filter(isActive=True).order_by('name')
        if productsubcategorys_obj.exists():
            page4 = self.paginate_queryset(productsubcategorys_obj)
            serializer=CustomProductSubCategorySerializers(page4,many=True)
            return self.get_paginated_response(serializer.data)
        else:
            return self.get_paginated_response([])


@api_view(['GET','POST'])
def product_subcategory_list(request):  
    subcategories_obj = ProductSubCategory.objects.filter(isActive=True).order_by('name')
    category_id=request.data.get('category_id')
    if category_id is not None and category_id !='':
        subcategories_obj=subcategories_obj.filter(category_id=category_id)
        
    if subcategories_obj.exists():
        serializer=ProductSubCategorySerializers(subcategories_obj,many=True)
        return Response({"response":{"n":1,"msg":"Product subcategory list fetched successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":1,"msg":"No product subcategories found","status":"success"},"data":[]})
    

@api_view(['POST'])
def add_product_subcategory(request):
    data = request.data.copy()
    data['isActive']=True

    alredy_exists = ProductSubCategory.objects.filter(Q(name__iexact=data.get('name')) & Q(isActive=True))
    if alredy_exists.exists():
        return Response({"response":{"n":0,"msg":"Product subcategory with this name already exists","status":"error"}})
    check_category = ProductCategory.objects.filter(id=data.get('category'), isActive=True)
    if not check_category.exists():
        return Response({"response":{"n":0,"msg":"Associated product category not found","status":"error"}})
    
    serializer = ProductSubCategorySerializers(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response({"response":{"n":1,"msg":"Product subcategory added successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":0,"msg":"Invalid data","status":"error"},"errors":serializer.errors})
    

@api_view(['POST'])
def update_product_subcategory(request):
    data = request.data.copy()
    subcategory_id = data.get('id')
    try:
        subcategory_obj = ProductSubCategory.objects.get(id=subcategory_id, isActive=True)
    except ProductSubCategory.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product subcategory not found","status":"error"}})
    
    alredy_exists = ProductSubCategory.objects.filter(Q(name__iexact=data.get('name')) & ~Q(id=subcategory_id) & Q(isActive=True))
    if alredy_exists.exists():
        return Response({"response":{"n":0,"msg":"Product subcategory with this name already exists","status":"error"}})
    check_category = ProductCategory.objects.filter(id=data.get('category'), isActive=True)
    if not check_category.exists():
        return Response({"response":{"n":0,"msg":"Associated product category not found","status":"error"}})
    
    serializer = ProductSubCategorySerializers(subcategory_obj, data=data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({"response":{"n":1,"msg":"Product subcategory updated successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":0,"msg":"Invalid data","status":"error"},"errors":serializer.errors})

@api_view(['POST'])
def delete_product_subcategory(request):
    data = request.data.copy()
    subcategory_id = data.get('id')
    try:
        subcategory_obj = ProductSubCategory.objects.get(id=subcategory_id, isActive=True)
    except ProductSubCategory.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product subcategory not found","status":"error"}})
    
    subcategory_obj.isActive = False
    subcategory_obj.save()
    return Response({"response":{"n":1,"msg":"Product subcategory deleted successfully","status":"success"}})
@api_view(['POST'])
def product_subcategory_by_id(request):

    data = request.data.copy()
    subcategory_id = data.get('id')
    try:
        subcategory_obj = ProductSubCategory.objects.get(id=subcategory_id, isActive=True)
    except ProductSubCategory.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product subcategory not found","status":"error"}})
    
    serializer=ProductSubCategorySerializers(subcategory_obj)
    return Response({"response":{"n":1,"msg":"Product subcategory fetched successfully","status":"success"},"data":serializer.data})



@api_view(['GET'])
def product_brand_list(request):
    brands_obj = ProductBrand.objects.filter(isActive=True).order_by('name')
    if brands_obj.exists():
        serializer=ProductBrandSerializers(brands_obj,many=True)
        return Response({"response":{"n":1,"msg":"Product brand list fetched successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":1,"msg":"No product brands found","status":"success"},"data":[]})
    
class product_brand_list_pagination(GenericAPIView):
    pagination_class = CustomPagination
    def post(self, request):
        brands_obj = ProductBrand.objects.filter(isActive=True).order_by('name')
        if brands_obj.exists():
            page4 = self.paginate_queryset(brands_obj)
            serializer=ProductBrandSerializers(page4,many=True)
            return self.get_paginated_response(serializer.data)
        else:
            return self.get_paginated_response([])

@api_view(['POST'])
def add_product_brand(request):
    data = request.data.copy()
    data['isActive']=True

    alredy_exists = ProductBrand.objects.filter(Q(name__iexact=data.get('name')) & Q(isActive=True))
    if alredy_exists.exists():
        return Response({"response":{"n":0,"msg":"Product brand with this name already exists","status":"error"}})
    
    serializer = ProductBrandSerializers(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response({"response":{"n":1,"msg":"Product brand added successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":0,"msg":"Invalid data","status":"error"},"errors":serializer.errors})
    
@api_view(['POST'])
def update_product_brand(request):
    data = request.data.copy()
    brand_id = data.get('id')
    try:
        brand_obj = ProductBrand.objects.get(id=brand_id, isActive=True)
    except ProductBrand.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product brand not found","status":"error"}})
    
    alredy_exists = ProductBrand.objects.filter(Q(name__iexact=data.get('name')) & ~Q(id=brand_id) & Q(isActive=True))
    if alredy_exists.exists():
        return Response({"response":{"n":0,"msg":"Product brand with this name already exists","status":"error"}})
    
    serializer = ProductBrandSerializers(brand_obj, data=data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({"response":{"n":1,"msg":"Product brand updated successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":0,"msg":"Invalid data","status":"error"},"errors":serializer.errors})
    
@api_view(['POST'])
def delete_product_brand(request):
    data = request.data.copy()
    brand_id = data.get('id')
    try:
        brand_obj = ProductBrand.objects.get(id=brand_id, isActive=True)
    except ProductBrand.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product brand not found","status":"error"}})
    
    brand_obj.isActive = False
    brand_obj.save()
    return Response({"response":{"n":1,"msg":"Product brand deleted successfully","status":"success"}})
@api_view(['POST'])
def product_brand_by_id(request):

    data = request.data.copy()
    brand_id = data.get('id')
    try:
        brand_obj = ProductBrand.objects.get(id=brand_id, isActive=True)
    except ProductBrand.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product brand not found","status":"error"}})
    
    serializer=ProductBrandSerializers(brand_obj)
    return Response({"response":{"n":1,"msg":"Product brand fetched successfully","status":"success"},"data":serializer.data})

class product_size_list_pagination(GenericAPIView):
    pagination_class = CustomPagination
    def post(self, request):
        productsizes_obj = ProductSizeUnit.objects.filter(isActive=True).order_by('name')
        if productsizes_obj.exists():
            page4 = self.paginate_queryset(productsizes_obj)
            serializer=ProductSizeUnitSerializers(page4,many=True)
            return self.get_paginated_response(serializer.data)
        else:
            return self.get_paginated_response([])
        




@api_view(['GET'])
def product_size_unit_list(request):
    sizes_obj = ProductSizeUnit.objects.filter(isActive=True).order_by('name')
    if sizes_obj.exists():
        serializer=ProductSizeUnitSerializers(sizes_obj,many=True)
        return Response({"response":{"n":1,"msg":"Product size list fetched successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":1,"msg":"No product sizes found","status":"success"},"data":[]})





@api_view(['POST'])
def add_product_size_unit(request):
    data = request.data.copy()
    data['isActive']=True

    alredy_exists = ProductSizeUnit.objects.filter(Q(name__iexact=data.get('name')) & Q(isActive=True))
    if alredy_exists.exists():
        return Response({"response":{"n":0,"msg":"Product size unit with this name already exists","status":"error"}})
    
    serializer = ProductSizeUnitSerializers(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response({"response":{"n":1,"msg":"Product size unit added successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":0,"msg":"Invalid data","status":"error"},"errors":serializer.errors})
    
@api_view(['POST'])
def update_product_size_unit(request):
    data = request.data.copy()
    size_unit_id = data.get('id')
    try:
        size_unit_obj = ProductSizeUnit.objects.get(id=size_unit_id, isActive=True)
    except ProductSizeUnit.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product size unit not found","status":"error"}})
    
    alredy_exists = ProductSizeUnit.objects.filter(Q(name__iexact=data.get('name')) & ~Q(id=size_unit_id) & Q(isActive=True))
    if alredy_exists.exists():
        return Response({"response":{"n":0,"msg":"Product size_unit with this name already exists","status":"error"}})
    
    serializer = ProductSizeUnitSerializers(size_unit_obj, data=data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({"response":{"n":1,"msg":"Product size_unit updated successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":0,"msg":"Invalid data","status":"error"},"errors":serializer.errors})
    
@api_view(['POST'])
def delete_product_size_unit(request):
    data = request.data.copy()
    size_unit_id = data.get('id')
    try:
        size_unit_obj = ProductSizeUnit.objects.get(id=size_unit_id, isActive=True)
    except ProductSizeUnit.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product size_unit not found","status":"error"}})
    
    size_unit_obj.isActive = False
    size_unit_obj.save()
    return Response({"response":{"n":1,"msg":"Product size_unit deleted successfully","status":"success"}})
@api_view(['POST'])
def product_size_unit_by_id(request):

    data = request.data.copy()
    size_unit_id = data.get('id')
    try:
        size_unit_obj = ProductSizeUnit.objects.get(id=size_unit_id, isActive=True)
    except ProductSizeUnit.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product size_unit not found","status":"error"}})
    
    serializer=ProductSizeUnitSerializers(size_unit_obj)
    return Response({"response":{"n":1,"msg":"Product size_unit fetched successfully","status":"success"},"data":serializer.data})























class product_color_list_pagination(GenericAPIView):
    pagination_class = CustomPagination
    def post(self, request):
        productcolors_obj = ProductColor.objects.filter(isActive=True).order_by('name')
        if productcolors_obj.exists():
            page4 = self.paginate_queryset(productcolors_obj)
            serializer=ProductColorSerializers(page4,many=True)
            return self.get_paginated_response(serializer.data)
        else:
            return self.get_paginated_response([])
@api_view(['GET'])
def product_color_list(request):
    colors_obj = ProductColor.objects.filter(isActive=True).order_by('name')
    if colors_obj.exists():
        serializer=ProductColorSerializers(colors_obj,many=True)
        return Response({"response":{"n":1,"msg":"Product color list fetched successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":1,"msg":"No product colors found","status":"success"},"data":[]})
    

@api_view(['POST'])
def add_product_color(request):
    data = request.data.copy()
    data['isActive']=True

    alredy_exists = ProductColor.objects.filter(Q(name__iexact=data.get('name')) & Q(isActive=True))
    if alredy_exists.exists():
        return Response({"response":{"n":0,"msg":"Product color with this name already exists","status":"error"}})
    
    serializer = ProductColorSerializers(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response({"response":{"n":1,"msg":"Product color added successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":0,"msg":"Invalid data","status":"error"},"errors":serializer.errors})
    
@api_view(['POST'])
def update_product_color(request):
    data = request.data.copy()
    color_id = data.get('id')
    try:
        color_obj = ProductColor.objects.get(id=color_id, isActive=True)
    except ProductColor.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product color not found","status":"error"}})
    
    alredy_exists = ProductColor.objects.filter(Q(name__iexact=data.get('name')) & ~Q(id=color_id) & Q(isActive=True))
    if alredy_exists.exists():
        return Response({"response":{"n":0,"msg":"Product color with this name already exists","status":"error"}})
    
    serializer = ProductColorSerializers(color_obj, data=data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({"response":{"n":1,"msg":"Product color updated successfully","status":"success"},"data":serializer.data})
    else:
        return Response({"response":{"n":0,"msg":"Invalid data","status":"error"},"errors":serializer.errors})
    
@api_view(['POST'])
def delete_product_color(request):
    data = request.data.copy()
    color_id = data.get('id')
    try:
        color_obj = ProductColor.objects.get(id=color_id, isActive=True)
    except ProductColor.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product color not found","status":"error"}})
    
    color_obj.isActive = False
    color_obj.save()
    return Response({"response":{"n":1,"msg":"Product color deleted successfully","status":"success"}})
@api_view(['POST'])
def product_color_by_id(request):

    data = request.data.copy()
    color_id = data.get('id')
    try:
        color_obj = ProductColor.objects.get(id=color_id, isActive=True)
    except ProductColor.DoesNotExist:
        return Response({"response":{"n":0,"msg":"Product color not found","status":"error"}})
    
    serializer=ProductColorSerializers(color_obj)
    return Response({"response":{"n":1,"msg":"Product color fetched successfully","status":"success"},"data":serializer.data})



import openpyxl
from io import BytesIO
from vendor.models import Vendor

@api_view(['POST'])
def bulk_upload_products_by_name(request):
    """
    Accepts an .xlsx file in form field 'file'.
    Excel must contain headers (first row) matching:
    name, price, category, subcategory, brand, color, size_unit, vendor, description, redirect_url
    All names are matched case-insensitive by .strip().lower()
    """

    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return Response({"response": {"n": 0, "msg": "No file uploaded"}}, status=400)

    # Basic file extension check
    if not uploaded_file.name.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        return Response({"response": {"n": 0, "msg": "Unsupported file type. Please upload .xlsx file"}}, status=400)

    try:
        wb = openpyxl.load_workbook(filename=BytesIO(uploaded_file.read()), data_only=True)
    except Exception as e:
        return Response({"response": {"n": 0, "msg": f"Failed to read Excel file: {str(e)}"}}, status=400)

    sheet = wb.active

    # Read header row and map column indices
    headers = {}
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, cell_value in enumerate(first_row):
        if cell_value is None:
            continue
        headers[str(cell_value).strip().lower()] = idx  # zero-based index

    # required header names
    expected_headers = ["name", "price", "category", "subcategory", "brand", "color", "size_unit", "vendor", "description", "redirect_url"]
    # Allow missing optional headers (description, redirect_url). But require name & price & category & brand & color & size_unit & vendor & subcategory
    missing_required = []
    for h in ["name", "price", "category", "subcategory", "brand", "color", "size_unit", "vendor"]:
        if h not in headers:
            missing_required.append(h)
    if missing_required:
        return Response({"response": {"n": 0, "msg": f"Missing required columns: {', '.join(missing_required)}. Headers must be: {', '.join(expected_headers)}"}}, status=400)

    # Helper to lookup by name (case-insensitive)
    def get_obj_by_name(model, name_value, extra_filter=None):
        if name_value is None:
            return None
        nm = str(name_value).strip()
        if nm == "":
            return None
        qs = model.objects.filter(name__iexact=nm)
        if extra_filter:
            qs = qs.filter(**extra_filter)
        return qs.first()

    created = 0
    errors = []
    rows_processed = 0

    # Start transaction: create all or only commit successful creates? We'll create all valid rows and report errors for invalid ones.
    # So we won't rollback all on single error — we create per-row if valid.
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        rows_processed += 1
        # read fields by header mapping (use .get to avoid KeyError)
        def val(h):
            idx = headers.get(h)
            if idx is None:
                return None
            return row[idx] if idx < len(row) else None

        row_name = val("name")
        row_price = val("price")
        row_category = val("category")
        row_subcategory = val("subcategory")
        row_brand = val("brand")
        row_color = val("color")
        row_size_unit = val("size_unit")
        row_vendor = val("vendor")
        row_description = val("description") if "description" in headers else None
        row_redirect = val("redirect_url") if "redirect_url" in headers else None

        # Basic validation
        row_errors = []

        if not row_name or str(row_name).strip() == "":
            row_errors.append("Missing product name")
        if row_price is None:
            row_errors.append("Missing price")
        else:
            try:
                # allow numeric strings too
                price_val = float(row_price)
                if price_val < 0:
                    row_errors.append("Price cannot be negative")
            except Exception:
                row_errors.append("Invalid price")

        # Resolve foreign keys by name
        category_obj = get_obj_by_name(ProductCategory, row_category)
        if not category_obj:
            row_errors.append(f"Category '{row_category}' not found")

        # Subcategory must be within category
        subcategory_obj = None
        if row_subcategory:
            # try to match subcategory name under that category if category_obj exists
            if category_obj:
                subcategory_obj = ProductSubCategory.objects.filter(category=category_obj, name__iexact=str(row_subcategory).strip()).first()
            if not subcategory_obj:
                # try global match if not found under category
                subcategory_obj = ProductSubCategory.objects.filter(name__iexact=str(row_subcategory).strip()).first()
            if not subcategory_obj:
                row_errors.append(f"Subcategory '{row_subcategory}' not found")

        brand_obj = get_obj_by_name(ProductBrand, row_brand)
        if not brand_obj:
            row_errors.append(f"Brand '{row_brand}' not found")

        color_obj = get_obj_by_name(ProductColor, row_color)
        if not color_obj:
            row_errors.append(f"Color '{row_color}' not found")

        size_obj = get_obj_by_name(ProductSizeUnit, row_size_unit)
        if not size_obj:
            row_errors.append(f"Size unit '{row_size_unit}' not found")

        vendor_obj = get_obj_by_name(Vendor, row_vendor)
        if not vendor_obj:
            row_errors.append(f"Vendor '{row_vendor}' not found")

        if row_errors:
            errors.append({"row": row_idx, "errors": row_errors})
            continue

        # All validations passed -> create product
        try:
            # Use atomic block per row to avoid partial DB state if something unexpected fails
            with transaction.atomic():
                product = Product.objects.create(
                    name=str(row_name).strip(),
                    price=float(row_price),
                    category=category_obj,
                    subcategory=subcategory_obj,
                    brand=brand_obj,
                    size_unit=size_obj,
                    color=color_obj,
                    vendor=vendor_obj,
                    description=str(row_description).strip() if row_description else None,
                    redirect_url=str(row_redirect).strip() if row_redirect else None
                )
                created += 1
        except Exception as e:
            errors.append({"row": row_idx, "errors": [f"DB error: {str(e)}"]})
            continue

    result = {
        "response": {"n": 1 if created>0 else 0, "msg": "Bulk upload finished"},
        "data": {
            "rows_processed": rows_processed,
            "created": created,
            "failed": len(errors),
            "errors": errors
        }
    }
    return Response(result)







